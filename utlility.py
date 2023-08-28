from selenium.webdriver.support.ui import WebDriverWait
import json
from selenium import webdriver
import tweet as TWEET
import csv
import re
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

class Utility:

    cell_names= ['A','B','C','D']

    def __init__(self):
        self.timeout = 100
        f = open('constants.json')
        data = json.load(f)
        self.configData = data

        f.close()
        pass



    def  get_last_index(self):
        workbook = load_workbook('twitter_comments.xlsx')
        sheet = workbook['Sheet1']
        return  len(sheet['A'])

    def write_tweet_to_excel(self, tweet: TWEET.Tweet):
        # We are using openpyxl library to Write tweet datat to Excel sheet.
        # A- Username
        # B - TweetText
        # C - Likes
        # D - Timestamp
        workbook = load_workbook('twitter_comments.xlsx')
        row_index = self.get_last_index() + 1 # To append data at the end of the excel file.
        worksheet = workbook.active
        tweet_as_list = tweet.get_tweet_as_list()  # this length of Tweet_as_list Must be equal to no. of columns in Excel.
        for i in  range(0,len(self.cell_names)):
            worksheet[f"{self.cell_names[i]}{row_index}"] = tweet_as_list[i]

        workbook.save('twitter_comments.xlsx')


    def write_to_csv(self, tweet:TWEET.Tweet):



        with open('students.csv', 'w', newline='') as file:
            writer = csv.writer(file)

            writer.writerow(tweet.get_tweet_as_list())
            file.close()

    def clear_excel(self):
        workbook = load_workbook('twitter_comments.xlsx')
        sheet = workbook['Sheet1']
        for i in range(0,len(sheet['A'])):
            sheet.delete_cols(1,7)




    def sign_in(self, driver):
        # element_present = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
        #                                                                                   self.configData[
        #                                                                                       "user_name_xpath"])))
        # # Wait for a maximum of 10 seconds until the element with the given selector is clickable
        # element_clickable = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,
        #                                                                                self.configData['user_name_xpath'])))
        # input_field_visible = WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH,
        #                                                                                        "user_name_xpath")))
        # Wait for a maximum of 10 seconds until the input field with the given selector is clickable
        input_field_interactable = WebDriverWait(driver, self.timeout).until(EC.element_to_be_clickable((By.XPATH,
                                                                                                         self.configData[
                                                                                                             "user_name_xpath"])))

        input_field_interactable.send_keys(self.configData['user_name'])  # Enter data into the input field
        ## STEP 2: Click on the next Button:

        next_button_clickable = WebDriverWait(driver, timeout=self.timeout).until(EC.element_to_be_clickable((By.XPATH,
                                                                                                              self.configData[
                                                                                                                  'next_button_xpath'])))
        next_button_clickable.click()

        ## STEP 3: Enter the pasword in password field.
        password_field_interactable = WebDriverWait(driver, self.timeout).until(EC.element_to_be_clickable((By.XPATH,
                                                                                                            self.configData[
                                                                                                                "password_field_xpath"])))
        password_field_interactable.send_keys(self.configData['password'])

        ## STEP 4: CLick on the login button:

        login_button_interactable = WebDriverWait(driver, timeout=self.timeout).until(
            EC.element_to_be_clickable((By.XPATH,
                                        self.configData['login_button_xpath'])))
        login_button_interactable.click()

    def search_feature(self, driver):

        search_field_interactable = WebDriverWait(driver, self.timeout).until(EC.element_to_be_clickable((By.XPATH,
                                                                                                          self.configData[
                                                                                                              "search_field_xpath"])))
        search_field_interactable.send_keys("NullCon Delhi")
        search_field_interactable.send_keys(Keys.RETURN)  # Enter data into the input field





    def open_user_handle(self, driver: webdriver.Chrome, link, login_flag = False):

        driver.get(link) # opens the page at given link to user handle
        print("STartting")
        # driver.execute_script("document.body.style.zoom='50%'")

#         LOGIN BUTTON AT THE BOTTOM:

        if login_flag:

            login_button_interactable = WebDriverWait(driver, timeout=self.timeout).until(
                EC.element_to_be_clickable((By.XPATH,
                                           "//*[@id='layers']/div/div[1]/div/div/div/div[2]/div[2]/div/div/div[1]/a")))
            login_button_interactable.click()

            self.sign_in(driver=driver)


        no_odf_scrolls = 30
        for i in range(0,no_odf_scrolls):
           print(f"SCROLL {i+1} of {no_odf_scrolls}")
           driver.implicitly_wait(5)
           div =  WebDriverWait(driver, self.timeout).until(
               EC.visibility_of_element_located((By.XPATH, "//article[@data-testid='tweet']")))
           print(f"this is the type of Var div: {type(div)}")
           articles = driver.find_elements(by=By.XPATH, value="//article[@data-testid='tweet']") #eg 4 article Tags

           print(f" This is the length of the tweet containers at this time: {len(articles)}")
           for article in articles: # iterate over each 4 and get data from each article tag.
             try:
                 print('-' * 40)
                 username = article.find_element(By.XPATH, ".//div[@data-testid='User-Name']").text
                 timeStamp = article.find_element(By.XPATH, ".//time").get_attribute('datetime')
                 tweet = article.find_element(By.XPATH, ".//div[@data-testid='tweetText']").text
                 reply = article.find_element(By.XPATH, ".//div[@data-testid='reply']").text
                 retweets = article.find_element(By.XPATH, ".//div[@data-testid='retweet']").text
                 likes = article.find_element(By.XPATH, ".//div[@data-testid='like']").text
                 # views = article.find_element(by=By.XPATH,value="//span[@class='css-901oao css-16my406 r-poiln3 r-bcqeeo r-qvutc0']").text

                 print(f" {tweet}")
                 print(f"By: {username}")
                 print(f"Replies: {reply}")

                 print(f"Time: {timeStamp}")

                 print(f"Likes: {likes}")

                 print(f"Retweets: {retweets}")

                 # print(f"Views: {views}")
                 print('-' * 40)
                 tweet_model = TWEET.Tweet(
                     userHandle=username,
                     tweetText=tweet,
                     timeStamp=timeStamp,
                     views=0,
                     bookmarks=0,
                     retweets=retweets,
                     quotes=0,
                     likes=likes
                 )
                 self.write_tweet_to_excel(tweet=tweet_model)

             except StaleElementReferenceException:
                 print("SKIPPING THE STALE ELEMENT REFERENCE AND PASSING")
                 pass
             except NoSuchElementException:
                 print("NOSUCH element exception found!")
                 pass
           driver.execute_script('window.scrollTo(0,document.body.scrollHeight);')
        # Close the WebDriver
        # driver.quit()

